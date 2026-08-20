# -*- coding: utf-8 -*-
"""Build formula-driven ECCO Kids mediaplan workbook (marmelad styling)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.marker import DataPoint as _  # noqa: F401 — keep import surface small

OUT = Path(__file__).with_name("ecco-kids-direct-mediaplan-formulas.xlsx")
OUT_ALT = Path(__file__).with_name("ecco-kids-direct-mediaplan-formulas-v2.xlsx")

NAVY = "1F4E79"
NAVY_FILL = PatternFill("solid", fgColor=NAVY)
CREAM = PatternFill("solid", fgColor="FCE4D6")
BLUE = PatternFill("solid", fgColor="D6EAF8")
GREEN = PatternFill("solid", fgColor="E8F5E9")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
LILAC = PatternFill("solid", fgColor="E8DAEF")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
STRIPE = PatternFill("solid", fgColor="F7FBFF")
TOTAL_FILL = PatternFill("solid", fgColor="D5F5E3")
WARN_FILL = PatternFill("solid", fgColor="FDEDEC")

FONT_WHITE = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=NAVY)
FONT_H2 = Font(name="Calibri", size=12, bold=True, color=NAVY)
FONT_BODY = Font(name="Calibri", size=11, color="333333")
FONT_MUTED = Font(name="Calibri", size=10, italic=True, color="555555")
FONT_SMALL = Font(name="Calibri", size=9, italic=True, color="555555")

THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
WRAP = Alignment(wrap_text=True, vertical="center")
LEFT = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

SHEET_PLAN_PARAMS = "Параметры плана"
SHEET_PLAN_PARAMS_REF = "'Параметры плана'"
PLAN_CELL_BUDGET = f"{SHEET_PLAN_PARAMS_REF}!$C$6"
PLAN_CELL_AOV = f"{SHEET_PLAN_PARAMS_REF}!$C$7"
PLAN_CELL_HORIZON = f"{SHEET_PLAN_PARAMS_REF}!$C$8"
PLAN_WEIGHT_TABLE = f"{SHEET_PLAN_PARAMS_REF}!$A$21:$B$32"
PLAN_WEIGHT_FIRST_ROW = 21
PLAN_WEIGHT_LAST_ROW = 32
PLAN_WEIGHT_TOTAL_ROW = 33

MONTHS = [
    ("2026-09", "Сен 2026", 9, 2026),
    ("2026-10", "Окт 2026", 10, 2026),
    ("2026-11", "Ноя 2026", 11, 2026),
    ("2026-12", "Дек 2026", 12, 2026),
    ("2027-01", "Янв 2027", 1, 2027),
    ("2027-02", "Фев 2027", 2, 2027),
    ("2027-03", "Мар 2027", 3, 2027),
    ("2027-04", "Апр 2027", 4, 2027),
    ("2027-05", "Май 2027", 5, 2027),
    ("2027-06", "Июн 2027", 6, 2027),
    ("2027-07", "Июл 2027", 7, 2027),
    ("2027-08", "Авг 2027", 8, 2027),
]

# (label col 2, weight) — sum = 12.0 → year total 12M при C6 = 1_000_000
MONTH_WEIGHTS: list[tuple[str, float]] = [
    ("Сен 2026", 1.15),
    ("Окт 2026", 1.10),
    ("Ноя 2026", 1.05),
    ("Дек 2026", 1.05),
    ("Янв 2027", 1.15),
    ("Фев 2027", 1.05),
    ("Мар 2027", 1.00),
    ("Апр 2027", 0.95),
    ("Май 2027", 0.90),
    ("Июн 2027", 0.85),
    ("Июл 2027", 0.80),
    ("Авг 2027", 0.95),
]

# name, тип кампании, тип площадки, доп. инструменты, тип объявлений, стратегия,
# share, ctr, cpc, cr, bench note
INSTRUMENTS = [
    (
        "ECCO Kids · бренд · Поиск",
        "ЕПК",
        "Поиск",
        "ключи бренд / бренд+товар",
        "текстово-графические",
        "Максимум конверсий",
        0.18, 0.10, 28, 0.065,
        "Бренд дешевле клика — задача забрать свой трафик, не растить охват.",
    ),
    (
        "ECCO Kids · категория boys · Поиск",
        "ЕПК",
        "Поиск",
        "ключи категория kids/boys/кроссовки",
        "текстово-графические",
        "Максимум конверсий",
        0.32, 0.045, 48, 0.038,
        "32% бюджета — основной спрос родителей по категории.",
    ),
    (
        "ECCO Kids · товарная галерея · Поиск",
        "ЕПК",
        "Поиск",
        "фид + товарные объявления",
        "товарные",
        "Максимум конверсий",
        0.15, 0.07, 44, 0.045,
        "Товарная галерея: CTR выше текста, CPC между брендом и категорией.",
    ),
    (
        "ECCO Kids · интересы и ретаргетинг · РСЯ",
        "ЕПК",
        "РСЯ",
        "интересы (kids/fashion/родители) + ретаргетинг визитов/корзины",
        "графические + текстово-графические",
        "Максимум конверсий",
        0.20, 0.008, 11, 0.025,
        "РСЯ с ручным таргетингом: CR выше автотаргета, ДРР контролируем.",
    ),
    (
        "ECCO Kids · автотаргетинг · РСЯ",
        "ЕПК",
        "РСЯ",
        "автотаргетинг",
        "графические + текстово-графические",
        "Максимум кликов",
        0.10, 0.008, 11, 0.015,
        "«Максимум кликов» — объём в сетях, CR ниже интересов.",
    ),
    (
        "ECCO Kids · защита бренда · РСЯ",
        "ЕПК",
        "РСЯ",
        "бренд в сетях",
        "графические + текстово-графические",
        "Максимум конверсий",
        0.05, 0.008, 11, 0.025,
        "Защита бренда в РСЯ — те же бенчи, что интересы+ретаргет.",
    ),
]

# ir = row on «Инструменты», platform, col D/E/F texts for split justification table
SPLIT_JUSTIFICATION: list[tuple[int, str, str, str, str]] = [
    (
        5,
        "Поиск",
        (
            "Тип: ЕПК, места показа только Поиск, объявления текстово-графические, "
            "стратегия «Максимум конверсий», цель Метрики «Покупка на сайте». "
            "Условие показа — ключи бренд и бренд+товар. "
            "Почему тип: бренд в Поиске — самый дешёвый и предсказуемый спрос. "
            "Смешивать с РСЯ нельзя: другой аукцион и другой CR."
        ),
        (
            "Почему 18% (180 000 ₽/мес): этого достаточно, чтобы забрать свой спрос, "
            "но не раздувать бренд сверх частотности. Основной объём категории должен жить в небренде."
        ),
        (
            "Экономика (допущение, не кабинет): CTR 10,0%, CPC 28 ₽ → ~6 429 кликов; "
            "CR 6,5% → ~418 покупок; CPA ≈ 431 ₽; выручка ≈ 1,78 млн ₽; ДРР ≈ 10%. "
            "Это якорь канала: низкий ДРР компенсирует более дорогой категорийный поиск."
        ),
    ),
    (
        6,
        "Поиск",
        (
            "Тип: ЕПК, Поиск, ТГО, «Максимум конверсий», ключи категория kids/boys/кроссовки/"
            "детская обувь — спрос родителей, не бренд. "
            "Почему тип: без категорийного Поиска нет объёма покупок в e-com обуви. "
            "Товарные объявления сюда не кладу: фид и семантика — разные рычаги оптимизации."
        ),
        (
            "Почему 32% (320 000 ₽) — крупнейшая доля: это основной небренд-спрос. "
            "Меньше 25% — недобор категории; больше 40% — ДРР канала вылетает из коридора, "
            "потому что CPC 48 ₽ и CR 3,8% тяжелее бренда."
        ),
        (
            "Экономика: CTR 4,5%, CPC 48 ₽ → ~6 667 кликов; CR 3,8% → ~253 покупки; "
            "CPA ≈ 1 265 ₽; ДРР строки ≈ 30%. Строка «дорогая» сознательно: "
            "её тянет бренд и РСЯ-ретаргет с ДРР ~10%."
        ),
    ),
    (
        7,
        "Поиск",
        (
            "Тип: ЕПК, Поиск, товарные объявления (фид), «Максимум конверсий». "
            "Не путать с типом кампании: ЕПК — кампания, товарное — тип объявления. "
            "Почему тип: карточка товара на Поиске поднимает CTR vs текст (7% vs 4,5%) "
            "и закрывает SKU kids/boys без отдельной «медийки»."
        ),
        (
            "Почему 15% (150 000 ₽): фид должен получить статистику на обучение стратегии, "
            "но не съесть семантику. 10% мало для фида, 25% дублирует категорийный поиск."
        ),
        (
            "Экономика: CTR 7,0%, CPC 44 ₽ → ~3 409 кликов; CR 4,5% → ~153 покупки; "
            "CPA ≈ 980 ₽; ДРР ≈ 23% — середина между брендом и категорией."
        ),
    ),
    (
        8,
        "РСЯ",
        (
            "Тип: ЕПК, места показа только РСЯ, графика + ТГО, «Максимум конверсий». "
            "Условие показа: интересы (kids/fashion/родители) + ретаргетинг визитов и корзины. "
            "Почему тип: сети закрывают тех, кто не ввёл запрос, и дожимают визит/корзину. "
            "Поиск сюда не мешаю."
        ),
        (
            "Почему 20% (200 000 ₽): e-com kids без ретаргета теряет догрев. "
            "Больше 25% РСЯ-интересов размывает ДРР: CTR сети 0,8%, конверсия холоднее Поиска."
        ),
        (
            "Экономика: CPC 11 ₽ → ~18 182 клика; CR 2,5% → ~455 покупок; "
            "CPA ≈ 440 ₽; ДРР ≈ 10%. Дешёвый клик + тёплый ретаргет — второй якорь ДРР после бренда."
        ),
    ),
    (
        9,
        "РСЯ",
        (
            "Тип: ЕПК, РСЯ, автотаргетинг, «Максимум кликов» "
            "(не «Максимум конверсий»: автотаргету сначала нужен объём кликов, CR ниже). "
            "Почему тип: автотаргетинг расширяет охват, но условие показа система берёт сама — "
            "CR падает до 1,5%. Держу отдельно, чтобы не портить статистику интересов/ретаргета."
        ),
        (
            "Почему 10% (100 000 ₽): тестовый объём. Резать первым, если факт ДРР > 25%. "
            "Не поднимать, пока CR не догонит интересы."
        ),
        (
            "Экономика: CPC 11 ₽ → ~9 091 клик; CR 1,5% → ~136 покупок; "
            "CPA ≈ 735 ₽; ДРР ≈ 17%. Плата за объём в сетях."
        ),
    ),
    (
        10,
        "РСЯ",
        (
            "Тип: ЕПК, РСЯ, бренд в сетях, «Максимум конверсий», графика + ТГО. "
            "Почему тип: защита бренда в РСЯ, чтобы конкурент не перехватывал показ по бренду "
            "вне Поиска. Не смешивать с бренд-Поиском."
        ),
        (
            "Почему 5% (50 000 ₽): частотность бренда в сетях ограничена. "
            "10%+ — переплата за те же аукционы."
        ),
        (
            "Экономика: CPC 11 ₽ → ~4 545 кликов; CR 2,5% → ~114 покупок; "
            "CPA ≈ 439 ₽; ДРР ≈ 10%."
        ),
    ),
]

FOOTER_ASSUMPTIONS = (
    "На листе разделены цифры, которые подставляются в формулы, и правила, "
    "по которым собран медиаплан. 1 000 000 ₽/мес без НДС — опорный месяц (вес 1,0); "
    "на сетке года к базовому расходу кампании умножаются сезонные веса из таблицы ниже — "
    "сумма весов 12, годовой расход канала 12 000 000 ₽. "
    "AOV 4 250 ₽ взяли с витрины kids/boys как рабочую оценку среднего чека: "
    "от него считается выручка и весь ДРР. Цель воронки — «Покупка на сайте»; "
    "KPI — ДРР в коридоре 15–25%. Пики сентябрь/октябрь (школа) и январь (зима) — "
    "усиливаем категорийный Поиск и товарную галерею, плюс интересы и ретаргетинг РСЯ; "
    "летом (июнь–июль) ниже базы — первым режем автотаргетинг РСЯ, бренд-Поиск и ретаргет не трогаем. "
    "Если после запуска средний чек или CR в кабинете расходятся с планом — "
    "правим параметры или бенчи на «Инструментах», а не пересчитываем вручную строки года."
)

FOOTER_INSTRUMENTS = (
    "Главное в сплите — два якоря с низким ДРР: бренд в Поиске и ретаргет в РСЯ. "
    "Они перекрывают «тяжёлую» категорийную строку с CPC 48 ₽ — без этого микс не проходит "
    "коридор 15–25%. Категория 32% — не жадность, а единственный источник объёма небренда; "
    "резать её первой нельзя. "
    "Автотаргетинг держу отдельно и с «Максимум кликов»: это буфер на объём, "
    "его снимаем при первом признаке просадки CR, не трогая бренд и ретаргет. "
    "Таблица обоснования выше — логика долей; цифры месяца пересчитаются сами "
    "при смене жёлтых бенчей."
)

FOOTER_YEAR_PLAN = (
    "Сезонность заложена через веса на «Параметрах плана»: сентябрь/октябрь и январь выше базы, "
    "июнь–июль ниже — годовой потолок 12 млн сохраняется. "
    "В пиках смещаем объём в категорийный Поиск, товарную галерею и ретаргет РСЯ; "
    "в просадке летом первым режем автотаргетинг РСЯ, бренд-Поиск и ретаргет не трогаем — "
    "они держат ДРР. "
    "При проверке смотрите не на каждую строку, а на разрыв: категорийный Поиск будет "
    "«краснее» по ДРР, чем бренд и интересы — так и задумано. "
    "Если после запуска факт по категории стабильно хуже бенча, правим CR/CPC на «Инструментах», "
    "а не пересчитываем веса месяцев вслепую."
)

FOOTER_MONTHLY = (
    "Месяцы различаются: SUMIF тянет расход с сетки года, где к базовому бюджету кампании "
    "умножен сезонный вес с «Параметров плана». "
    "ДРР месяца считаю от сумм, не усредняю ДРР кампаний: иначе категория с ДРР ~30% "
    "«спрячется» за брендом. "
    "Сентябрь/октябрь и январь выше июня/июля — это школьный и зимний спрос; "
    "сумма 12 месяцев = 12 млн ₽, веса и логика сдвига описаны на «Параметрах плана»."
)

FOOTER_YEAR_SUMMARY = (
    "Приёмка плана — годовой ДРР в коридоре 15–25% (строка проверки). "
    "Годовые KPI — сумма 12 сезонных месяцев с «Медиаплана на год», не flat × 12. "
    "По вкладу в покупки лидирует не «самая дешёвая» кампания, а связка: "
    "категорийный Поиск даёт объём, ретаргет РСЯ — догрев, бренд — дешёвый захват своего спроса. "
    "Автотаргетинг с худшим CR не ломает год, потому что доля 10% и изолирована. "
    "После старта сверяем факт CR по каждому месту показа; "
    "бенчи в плане — рабочая гипотеза, не статистика кабинета ECCO."
)


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = NAVY_FILL
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN


def apply_body(cell, fill=None, fmt=None, align=None) -> None:
    cell.font = FONT_BODY
    cell.border = THIN
    cell.alignment = align or WRAP
    if fill is not None:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def footer(ws, row: int, text: str, cols: int = 8, body_rows: int = 4) -> None:
    """Write calculation comment footer with 4–6 merged body rows (~120 pt total height)."""
    body_rows = max(4, min(6, body_rows))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.cell(row, 1, "Комментарий к расчёту").font = FONT_H2
    ws.cell(row, 1).alignment = LEFT
    end_body = row + body_rows
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=end_body, end_column=cols)
    cell = ws.cell(row + 1, 1, text)
    cell.font = FONT_BODY
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    per_row = max(24, 120 // body_rows)
    for r in range(row + 1, end_body + 1):
        ws.row_dimensions[r].height = per_row


def build_split_justification_table(ws, start_row: int) -> int:
    """Insert «Обоснование сплита и расчёта» table; return footer start row."""
    ws.cell(start_row, 1, "Обоснование сплита и расчёта").font = FONT_H2
    hdr = start_row + 1
    split_headers = [
        "Кампания",
        "Места показа",
        "Доля / расход",
        "Почему этот тип кампании и условие показа",
        "Почему такая доля бюджета",
        "Экономика месяца",
    ]
    for i, h in enumerate(split_headers, 1):
        ws.cell(hdr, i, h)
    style_header_row(ws, hdr, 6)
    ws.row_dimensions[hdr].height = 36

    for idx, (ir, platform, col_d, col_e, col_f) in enumerate(SPLIT_JUSTIFICATION):
        r = hdr + 1 + idx
        fill = WHITE_FILL if idx % 2 else STRIPE
        ws.cell(r, 1, f"=B{ir}")
        ws.cell(r, 2, platform)
        ws.cell(r, 3, f'=TEXT(H{ir},"0%")&" ("&TEXT(I{ir},"#,##0")&" ₽/мес)"')
        ws.cell(r, 4, col_d)
        ws.cell(r, 5, col_e)
        ws.cell(r, 6, col_f)
        for c in range(1, 7):
            apply_body(ws.cell(r, c), fill=fill, align=LEFT)
        ws.row_dimensions[r].height = 80

    return hdr + len(SPLIT_JUSTIFICATION) + 2


def save_workbook(wb: Workbook) -> Path:
    """Save workbook; fall back to v2 filename on PermissionError."""
    OUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT)
        return OUT
    except PermissionError:
        wb.save(OUT_ALT)
        return OUT_ALT


def verify_justification(path: Path) -> tuple[int, int, int]:
    """Verify split table and long footer exist on «Инструменты». Returns row refs."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False)
    wi = wb["Инструменты"]
    title_row = None
    header_row = None
    last_data = None
    for r in range(1, 40):
        val = wi.cell(r, 1).value
        if val and "Обоснование сплита" in str(val):
            title_row = r
            header_row = r + 1
            last_data = r + 7
            break
    if title_row is None:
        raise RuntimeError("justification table title not found")
    if wi.cell(header_row, 1).value != "Кампания":
        raise RuntimeError("justification header row missing")

    combined_len = 0
    for r in range(header_row, last_data + 1):
        for c in range(1, 7):
            val = wi.cell(r, c).value
            if val:
                combined_len += len(str(val))
    for r in range(last_data + 1, last_data + 12):
        val = wi.cell(r, 1).value
        if val:
            combined_len += len(str(val))
    if combined_len < 800:
        raise RuntimeError(f"justification+footer too short: {combined_len} chars")
    return title_row, header_row, last_data


def summarize_seasonality() -> dict[str, float | str]:
    """Compute expected seasonality metrics from model constants (formula-equivalent)."""
    base_budget = 1_000_000
    aov = 4250
    weight_by_label = {label: w for label, w in MONTH_WEIGHTS}
    month_spend: dict[str, float] = {}
    year_spend = 0.0
    year_revenue = 0.0
    for _code, label, _m, _y in MONTHS:
        weight = weight_by_label[label]
        channel = round(base_budget * weight, 0)
        month_spend[label] = channel
        for _name, *_rest, share, ctr, cpc, cr, _note in INSTRUMENTS:
            budget = round(base_budget * share * weight, 0)
            clicks = round(budget / cpc, 0) if cpc else 0
            purchases = round(clicks * cr, 0)
            revenue = round(purchases * aov, 0)
            year_spend += budget
            year_revenue += revenue
    sep = month_spend["Сен 2026"]
    jul = month_spend["Июл 2027"]
    drr = year_spend / year_revenue if year_revenue else 0.0
    return {
        "weight_table": f"A{PLAN_WEIGHT_FIRST_ROW}:B{PLAN_WEIGHT_LAST_ROW}",
        "sep_spend": sep,
        "jul_spend": jul,
        "year_spend": year_spend,
        "year_drr_pct": drr * 100,
    }


def build() -> None:
    wb = Workbook()

    # ── Параметры плана ────────────────────────────────────────
    ws = wb.active
    ws.title = SHEET_PLAN_PARAMS
    ws.sheet_properties.tabColor = NAVY
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = "landscape"
    ws.print_title_rows = "1:5"
    ws.oddHeader.left.text = "ECCO Kids · Яндекс Директ"
    ws.oddFooter.right.text = "Лист «Параметры плана»"

    ws.merge_cells("A1:D1")
    ws["A1"] = "Параметры плана: ECCO Kids — медиаплан Яндекс Директ"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A2:D2")
    ws["A2"] = (
        "Бюджет и AOV — цифры для формул; ниже — условия, "
        "по которым собран сплит кампаний и воронка до покупки."
    )
    ws["A2"].font = FONT_MUTED
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 32

    ws.cell(4, 1, "Данные для расчёта").font = FONT_H2
    data_headers = ["Источник", "Параметр", "Значение", "Ед."]
    for i, h in enumerate(data_headers, 1):
        ws.cell(5, i, h)
    style_header_row(ws, 5, 4)

    data_rows = [
        ("из исходных данных", "Месячный бюджет без НДС", 1_000_000, "₽", YELLOW),
        ("информация с сайта", "Средний чек (AOV)", 4250, "₽", YELLOW),
        ("из исходных данных", "Горизонт планирования", 12, "мес", CREAM),
    ]
    for idx, (src, param, val, unit, fill) in enumerate(data_rows, 6):
        ws.cell(idx, 1, src)
        ws.cell(idx, 2, param)
        ws.cell(idx, 3, val)
        ws.cell(idx, 4, unit)
        for c in range(1, 5):
            apply_body(ws.cell(idx, c), fill=fill)

    ws["C6"].number_format = '#,##0" ₽"'
    ws["C7"].number_format = '#,##0" ₽"'
    ws["C8"].number_format = "0"

    ws["C6"].comment = Comment(
        f"Вход модели. Лист «Инструменты» берёт бюджет как {PLAN_CELL_BUDGET} × доля.",
        "model",
    )
    ws["C7"].comment = Comment("Вход модели. Выручка = покупки × этот AOV.", "model")

    ws.cell(10, 1, "Условия расчёта").font = FONT_H2
    cond_headers = ["Источник", "Условие", "Значение"]
    for i, h in enumerate(cond_headers, 1):
        ws.cell(11, i, h)
    style_header_row(ws, 11, 3)

    cond_rows = [
        ("из исходных данных", "Целевое действие", "Покупка на сайте (Метрика)"),
        ("из исходных данных", "KPI эффективности", "ДРР = расход без НДС / выручка"),
        ("из исходных данных", "Коридор ДРР", "15–25%"),
        ("принимаем значение", "Распределение бюджета по месяцам", (
            "Сезонные веса (таблица ниже): пики сен/окт и янв; лето ниже базы; "
            "год = 12 млн ₽ без НДС"
        )),
        ("принимаем значение", "Места показа", "Поиск и РСЯ — отдельные ЕПК, не смешиваем"),
        ("принимаем значение", "Воронка", "Показы → клики → покупки → выручка"),
    ]
    for idx, (src, cond, val) in enumerate(cond_rows, 12):
        fill = WHITE_FILL if (idx - 12) % 2 else STRIPE
        ws.cell(idx, 1, src)
        ws.cell(idx, 2, cond)
        ws.cell(idx, 3, val)
        for c in range(1, 4):
            apply_body(ws.cell(idx, c), fill=fill)

    ws.cell(19, 1, "Сезонность по месяцам").font = FONT_H2
    weight_headers = ["Месяц", "Вес", "Бюджет канала, ₽"]
    for i, h in enumerate(weight_headers, 1):
        ws.cell(20, i, h)
    style_header_row(ws, 20, 3)

    for idx, (label, weight) in enumerate(MONTH_WEIGHTS):
        r = PLAN_WEIGHT_FIRST_ROW + idx
        fill = WHITE_FILL if idx % 2 else STRIPE
        ws.cell(r, 1, label)
        ws.cell(r, 2, weight)
        ws.cell(r, 3, f"=ROUND($C$6*B{r},0)")
        for c in range(1, 4):
            apply_body(ws.cell(r, c), fill=fill, align=CENTER if c != 1 else LEFT)
        ws.cell(r, 2).number_format = "0.00"
        ws.cell(r, 3).number_format = '#,##0" ₽"'
        ws.row_dimensions[r].height = 22

    ws.cell(PLAN_WEIGHT_TOTAL_ROW, 1, "Итого")
    ws.cell(PLAN_WEIGHT_TOTAL_ROW, 2, f"=SUM(B{PLAN_WEIGHT_FIRST_ROW}:B{PLAN_WEIGHT_LAST_ROW})")
    ws.cell(
        PLAN_WEIGHT_TOTAL_ROW,
        3,
        f"=SUM(C{PLAN_WEIGHT_FIRST_ROW}:C{PLAN_WEIGHT_LAST_ROW})",
    )
    for c in range(1, 4):
        apply_body(ws.cell(PLAN_WEIGHT_TOTAL_ROW, c), fill=TOTAL_FILL)
        ws.cell(PLAN_WEIGHT_TOTAL_ROW, c).font = Font(
            name="Calibri", size=11, bold=True, color=NAVY
        )
    ws.cell(PLAN_WEIGHT_TOTAL_ROW, 2).number_format = "0.00"
    ws.cell(PLAN_WEIGHT_TOTAL_ROW, 3).number_format = '#,##0" ₽"'

    plan_footer_row = PLAN_WEIGHT_TOTAL_ROW + 2
    footer(
        ws, plan_footer_row,
        FOOTER_ASSUMPTIONS,
        4,
        body_rows=5,
    )
    set_widths(ws, {"A": 22, "B": 32, "C": 36, "D": 10})
    ws.auto_filter.ref = "A5:D8"
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[11].height = 22
    for r in range(6, 9):
        ws.row_dimensions[r].height = 22
    for r in range(12, 18):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[20].height = 22

    # ── Инструменты ────────────────────────────────────────────
    wi = wb.create_sheet("Инструменты")
    wi.sheet_properties.tabColor = "2E86C1"
    wi.sheet_view.showGridLines = False
    wi.freeze_panes = "A5"
    wi.page_setup.orientation = "landscape"
    wi.page_setup.fitToPage = True
    wi.page_setup.fitToWidth = 1
    wi.print_title_rows = "1:4"

    wi.merge_cells("A1:N1")
    wi["A1"] = "Инструменты: 6 кампаний ЕПК и бенчмарки воронки"
    wi["A1"].font = FONT_TITLE
    wi.merge_cells("A2:N2")
    wi["A2"] = (
        "Поиск и РСЯ в одной кампании не смешиваю: иначе не прочитаем аукцион и CR. "
        "32% на категорийный поиск — основной спрос родителей. "
        "Бренд дешевле клика, его задача — забрать свой трафик, не растить охват. "
        f"Доля × бюджет из листа «{SHEET_PLAN_PARAMS}» ({PLAN_CELL_BUDGET}) — "
        "опорный месяц (вес 1,0); на сетке года к расходу кампании умножаются сезонные веса. "
        "CTR/CPC/CR — жёлтые входы."
    )
    wi["A2"].font = FONT_MUTED
    wi.row_dimensions[2].height = 40

    ih = [
        "#", "Кампания", "Тип кампании", "Тип площадки", "Доп. инструменты",
        "Тип объявлений", "Стратегия", "Доля бюджета", "Бюджет/мес, ₽",
        "CTR", "CPC, ₽", "CR в покупку", "Клики/мес", "Покупки/мес",
    ]
    for i, h in enumerate(ih, 1):
        wi.cell(4, i, h)
    style_header_row(wi, 4, 14)

    platform_dv = DataValidation(
        type="list",
        formula1='"Поиск,РСЯ,Поиск и РСЯ"',
        allow_blank=False,
    )
    platform_dv.error = "Выберите место показа из списка"
    platform_dv.errorTitle = "Тип площадки"
    wi.add_data_validation(platform_dv)

    for i, (name, ctype, platform, extra, ads, strategy, share, ctr, cpc, cr, note) in enumerate(
        INSTRUMENTS, 1
    ):
        r = 4 + i
        fill = WHITE_FILL if i % 2 else STRIPE
        wi.cell(r, 1, i)
        wi.cell(r, 2, name)
        wi.cell(r, 3, ctype)
        wi.cell(r, 4, platform)
        wi.cell(r, 5, extra)
        wi.cell(r, 6, ads)
        wi.cell(r, 7, strategy)
        wi.cell(r, 8, share)
        wi.cell(r, 9, f"=ROUND({PLAN_CELL_BUDGET}*H{r},0)")
        wi.cell(r, 10, ctr)
        wi.cell(r, 11, cpc)
        wi.cell(r, 12, cr)
        wi.cell(r, 13, f"=IF(K{r}=0,0,ROUND(I{r}/K{r},0))")
        wi.cell(r, 14, f"=ROUND(M{r}*L{r},0)")
        for c in range(1, 15):
            apply_body(wi.cell(r, c), fill=fill, align=CENTER if c != 2 else LEFT)
        wi.cell(r, 2).alignment = LEFT
        wi.cell(r, 5).alignment = LEFT
        wi.cell(r, 8).number_format = "0.0%"
        wi.cell(r, 9).number_format = '#,##0'
        wi.cell(r, 10).number_format = "0.00%"
        wi.cell(r, 11).number_format = '#,##0'
        wi.cell(r, 12).number_format = "0.00%"
        wi.cell(r, 13).number_format = '#,##0'
        wi.cell(r, 14).number_format = '#,##0'
        for col in (8, 10, 11, 12):
            wi.cell(r, col).fill = YELLOW
        wi.cell(r, 2).comment = Comment(note, "bench")
        platform_dv.add(wi.cell(r, 4))

    # total row 11
    wi.cell(11, 1, "")
    wi.cell(11, 2, "Итого / месяц")
    wi.cell(11, 8, "=SUM(H5:H10)")
    wi.cell(11, 9, "=SUM(I5:I10)")
    wi.cell(11, 13, "=SUM(M5:M10)")
    wi.cell(11, 14, "=SUM(N5:N10)")
    for c in range(1, 15):
        apply_body(wi.cell(11, c), fill=TOTAL_FILL)
        wi.cell(11, c).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    wi["H11"].number_format = "0.0%"
    wi["I11"].number_format = '#,##0'
    wi["M11"].number_format = '#,##0'
    wi["N11"].number_format = '#,##0'

    # derived monthly KPI
    wi["A13"] = "Проверка месяца (формулы)"
    wi["A13"].font = FONT_H2
    wi.merge_cells("A14:B14")
    wi["A14"] = "Выручка / мес"
    wi["C14"] = f"=ROUND(N11*{PLAN_CELL_AOV},0)"
    wi["D14"] = "₽"
    wi.merge_cells("A15:B15")
    wi["A15"] = "ДРР / мес"
    wi["C15"] = '=IF(C14=0,0,I11/C14)'
    wi["D15"] = "%"
    wi.merge_cells("A16:B16")
    wi["A16"] = "CPA / мес"
    wi["C16"] = '=IF(N11=0,0,I11/N11)'
    wi["D16"] = "₽"
    wi.merge_cells("A17:B17")
    wi["A17"] = "Контроль долей"
    wi["C17"] = '=IF(ABS(H11-1)<0.0001,"OK: 100%","Ошибка: сумма долей ≠ 100%")'
    for r in range(14, 18):
        for c in range(1, 5):
            apply_body(wi.cell(r, c), fill=GREEN if r < 17 else CREAM)
    wi["C14"].number_format = '#,##0'
    wi["C15"].number_format = "0.0%"
    wi["C16"].number_format = '#,##0.0'
    wi["C14"].fill = YELLOW
    wi["C15"].fill = YELLOW
    wi["C16"].fill = YELLOW

    split_footer_row = build_split_justification_table(wi, 19)
    footer(wi, split_footer_row, FOOTER_INSTRUMENTS, 14, body_rows=5)
    set_widths(
        wi,
        {
            "A": 6, "B": 34, "C": 14, "D": 28, "E": 38, "F": 38, "G": 22,
            "H": 14, "I": 16, "J": 10, "K": 12, "L": 14, "M": 12, "N": 14,
        },
    )
    wi.auto_filter.ref = "A4:N10"
    wi.row_dimensions[4].height = 36
    for r in range(5, 12):
        wi.row_dimensions[r].height = 24

    # ── Медиаплан на год ───────────────────────────────────────
    wy = wb.create_sheet("Медиаплан на год")
    wy.sheet_properties.tabColor = "1ABC9C"
    wy.sheet_view.showGridLines = False
    wy.freeze_panes = "D5"
    wy.page_setup.orientation = "landscape"
    wy.page_setup.fitToPage = True
    wy.page_setup.fitToWidth = 1
    wy.page_setup.paperSize = wy.PAPERSIZE_A4
    wy.print_title_rows = "1:4"
    wy.print_title_cols = "A:C"

    wy.merge_cells("A1:N1")
    wy["A1"] = "Медиаплан на год: 6 кампаний × 12 месяцев (все цифры — формулы)"
    wy["A1"].font = FONT_TITLE
    wy.merge_cells("A2:N2")
    wy["A2"] = (
        "Сетка — 6 ЕПК × 12 месяцев; бюджет строки = база «Инструментов» × сезонный вес "
        f"с «{SHEET_PLAN_PARAMS}» ({PLAN_WEIGHT_TABLE}). "
        "ДРР года = сумма расхода / сумма выручки, не среднее ДРР строк. "
        "Крутим жёлтые CTR/CPC/CR на «Инструментах», не копируем итоги. "
        "Показы = клики / CTR; клики = бюджет / CPC; покупки = клики × CR; выручка = покупки × AOV."
    )
    wy["A2"].font = FONT_MUTED
    wy.row_dimensions[2].height = 40

    yh = [
        "Месяц", "Кампания", "Тип площадки", "Доп. инструменты", "Стратегия",
        "Бюджет, ₽", "CTR", "CPC, ₽", "Показы", "Клики", "CR", "Покупки",
        "Выручка, ₽", "ДРР",
    ]
    for i, h in enumerate(yh, 1):
        wy.cell(4, i, h)
    style_header_row(wy, 4, 14)

    r = 5
    for mi, (code, label, _m, _y) in enumerate(MONTHS):
        for ii in range(6):
            ir = 5 + ii
            fill = BLUE if mi % 2 == 0 else WHITE_FILL
            wy.cell(r, 1, label)
            wy.cell(r, 2, f"=Инструменты!B{ir}")
            wy.cell(r, 3, f"=Инструменты!D{ir}")
            wy.cell(r, 4, f"=Инструменты!E{ir}")
            wy.cell(r, 5, f"=Инструменты!G{ir}")
            wy.cell(r, 6, (
                f"=ROUND(Инструменты!I{ir}*VLOOKUP(A{r},{PLAN_WEIGHT_TABLE},2,FALSE),0)"
            ))
            wy.cell(r, 7, f"=Инструменты!J{ir}")
            wy.cell(r, 8, f"=Инструменты!K{ir}")
            wy.cell(r, 10, f"=IF(H{r}=0,0,ROUND(F{r}/H{r},0))")
            wy.cell(r, 9, f"=IF(G{r}=0,0,ROUND(J{r}/G{r},0))")
            wy.cell(r, 11, f"=Инструменты!L{ir}")
            wy.cell(r, 12, f"=ROUND(J{r}*K{r},0)")
            wy.cell(r, 13, f"=ROUND(L{r}*{PLAN_CELL_AOV},0)")
            wy.cell(r, 14, f"=IF(M{r}=0,0,F{r}/M{r})")
            for c in range(1, 15):
                apply_body(wy.cell(r, c), fill=fill, align=CENTER if c not in (2, 4) else LEFT)
            wy.cell(r, 6).number_format = '#,##0'
            wy.cell(r, 7).number_format = "0.00%"
            wy.cell(r, 8).number_format = '#,##0'
            wy.cell(r, 9).number_format = '#,##0'
            wy.cell(r, 10).number_format = '#,##0'
            wy.cell(r, 11).number_format = "0.00%"
            wy.cell(r, 12).number_format = '#,##0'
            wy.cell(r, 13).number_format = '#,##0'
            wy.cell(r, 14).number_format = "0.0%"
            r += 1

    last_data = r - 1
    wy.cell(r, 1, "")
    wy.cell(r, 2, "Итого за год")
    wy.cell(r, 6, f"=SUM(F5:F{last_data})")
    wy.cell(r, 9, f"=SUM(I5:I{last_data})")
    wy.cell(r, 10, f"=SUM(J5:J{last_data})")
    wy.cell(r, 12, f"=SUM(L5:L{last_data})")
    wy.cell(r, 13, f"=SUM(M5:M{last_data})")
    wy.cell(r, 14, f"=IF(M{r}=0,0,F{r}/M{r})")
    for c in range(1, 15):
        apply_body(wy.cell(r, c), fill=TOTAL_FILL)
        wy.cell(r, c).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    wy.cell(r, 6).number_format = '#,##0'
    wy.cell(r, 9).number_format = '#,##0'
    wy.cell(r, 10).number_format = '#,##0'
    wy.cell(r, 12).number_format = '#,##0'
    wy.cell(r, 13).number_format = '#,##0'
    wy.cell(r, 14).number_format = "0.0%"
    total_row = r

    wy.conditional_formatting.add(
        f"N5:N{last_data}",
        ColorScaleRule(
            start_type="num", start_value=0.12, start_color="27AE60",
            mid_type="num", mid_value=0.18, mid_color="F7DC6F",
            end_type="num", end_value=0.28, end_color="E74C3C",
        ),
    )

    footer(
        wy, total_row + 2,
        FOOTER_YEAR_PLAN,
        14,
        body_rows=5,
    )
    set_widths(
        wy,
        {
            "A": 12, "B": 34, "C": 14, "D": 28, "E": 20, "F": 14, "G": 10,
            "H": 10, "I": 12, "J": 10, "K": 10, "L": 11, "M": 14, "N": 10,
        },
    )
    wy.auto_filter.ref = f"A4:N{last_data}"
    wy.row_dimensions[4].height = 32

    # ── Сводка по месяцам ──────────────────────────────────────
    wm = wb.create_sheet("Сводка по месяцам")
    wm.sheet_properties.tabColor = "8E44AD"
    wm.sheet_view.showGridLines = False
    wm.freeze_panes = "A5"
    wm.page_setup.orientation = "landscape"
    wm.page_setup.fitToPage = True
    wm.page_setup.fitToWidth = 1
    wm.print_title_rows = "1:4"

    wm.merge_cells("A1:H1")
    wm["A1"] = "Сводка по месяцам: SUMIF с листа «Медиаплан на год»"
    wm["A1"].font = FONT_TITLE
    wm.merge_cells("A2:H2")
    wm["A2"] = (
        "SUMIF с листа «Медиаплан на год»: месяцы различаются по сезонным весам "
        f"с «{SHEET_PLAN_PARAMS}». "
        "ДРР месяца = сумма расхода / сумма выручки, не среднее ДРР кампаний."
    )
    wm["A2"].font = FONT_MUTED

    mh = ["Месяц", "Расход, ₽", "Показы", "Клики", "Покупки", "Выручка, ₽", "ДРР", "CPA, ₽"]
    for i, h in enumerate(mh, 1):
        wm.cell(4, i, h)
    style_header_row(wm, 4, 8)

    plan = "'Медиаплан на год'"
    for i, (_code, label, _m, _y) in enumerate(MONTHS):
        r = 5 + i
        fill = WHITE_FILL if i % 2 else STRIPE
        wm.cell(r, 1, label)
        wm.cell(r, 2, f"=SUMIF({plan}!A:A,A{r},{plan}!F:F)")
        wm.cell(r, 3, f"=SUMIF({plan}!A:A,A{r},{plan}!I:I)")
        wm.cell(r, 4, f"=SUMIF({plan}!A:A,A{r},{plan}!J:J)")
        wm.cell(r, 5, f"=SUMIF({plan}!A:A,A{r},{plan}!L:L)")
        wm.cell(r, 6, f"=SUMIF({plan}!A:A,A{r},{plan}!M:M)")
        wm.cell(r, 7, f"=IF(F{r}=0,0,B{r}/F{r})")
        wm.cell(r, 8, f"=IF(E{r}=0,0,B{r}/E{r})")
        for c in range(1, 9):
            apply_body(wm.cell(r, c), fill=fill)
        wm.cell(r, 2).number_format = '#,##0'
        wm.cell(r, 3).number_format = '#,##0'
        wm.cell(r, 4).number_format = '#,##0'
        wm.cell(r, 5).number_format = '#,##0'
        wm.cell(r, 6).number_format = '#,##0'
        wm.cell(r, 7).number_format = "0.0%"
        wm.cell(r, 8).number_format = '#,##0.0'

    wm.cell(17, 1, "Итого год")
    wm.cell(17, 2, "=SUM(B5:B16)")
    wm.cell(17, 3, "=SUM(C5:C16)")
    wm.cell(17, 4, "=SUM(D5:D16)")
    wm.cell(17, 5, "=SUM(E5:E16)")
    wm.cell(17, 6, "=SUM(F5:F16)")
    wm.cell(17, 7, "=IF(F17=0,0,B17/F17)")
    wm.cell(17, 8, "=IF(E17=0,0,B17/E17)")
    for c in range(1, 9):
        apply_body(wm.cell(17, c), fill=TOTAL_FILL)
        wm.cell(17, c).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    wm["B17"].number_format = '#,##0'
    wm["C17"].number_format = '#,##0'
    wm["D17"].number_format = '#,##0'
    wm["E17"].number_format = '#,##0'
    wm["F17"].number_format = '#,##0'
    wm["G17"].number_format = "0.0%"
    wm["H17"].number_format = '#,##0.0'

    wm.conditional_formatting.add(
        "G5:G16",
        ColorScaleRule(
            start_type="num", start_value=0.12, start_color="27AE60",
            mid_type="num", mid_value=0.18, mid_color="F7DC6F",
            end_type="num", end_value=0.28, end_color="E74C3C",
        ),
    )

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Расход и выручка по месяцам"
    chart.y_axis.title = "₽"
    chart.x_axis.title = None
    data = Reference(wm, min_col=2, min_row=4, max_col=2, max_row=16)
    data2 = Reference(wm, min_col=6, min_row=4, max_col=6, max_row=16)
    cats = Reference(wm, min_col=1, min_row=5, max_row=16)
    chart.add_data(data, titles_from_data=True)
    chart.add_data(data2, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.style = 10
    chart.y_axis.numFmt = '#,##0'
    chart.legend.position = "b"
    chart.width = 18
    chart.height = 8
    wm.add_chart(chart, "A28")

    footer(
        wm, 19,
        FOOTER_MONTHLY,
        8,
        body_rows=4,
    )
    set_widths(wm, {"A": 14, "B": 14, "C": 14, "D": 12, "E": 12, "F": 14, "G": 10, "H": 12})
    wm.auto_filter.ref = "A4:H16"

    # ── Сводка за год ──────────────────────────────────────────
    wg = wb.create_sheet("Сводка за год")
    wg.sheet_properties.tabColor = "C0392B"
    wg.sheet_view.showGridLines = False
    wg.freeze_panes = "A5"
    wg.page_setup.orientation = "landscape"
    wg.page_setup.fitToPage = True
    wg.page_setup.fitToWidth = 1
    wg.print_title_rows = "1:4"

    wg.merge_cells("A1:K1")
    wg["A1"] = "Сводка за год: инструменты и KPI (SUMIF с «Медиаплан на год»)"
    wg["A1"].font = FONT_TITLE
    wg.merge_cells("A2:K2")
    wg["A2"] = (
        "Коридор ДРР 15–25% — условие приёмки плана. "
        "Годовые KPI — сумма 12 сезонных месяцев с сетки года; доли и бенчи не меняются."
    )
    wg["A2"].font = FONT_MUTED

    gh = [
        "Кампания", "Тип площадки", "Доп. инструменты", "Доля", "Расход год, ₽",
        "Клики год", "Покупки год", "Выручка год, ₽", "ДРР", "CPA, ₽", "Вклад в покупки",
    ]
    for i, h in enumerate(gh, 1):
        wg.cell(4, i, h)
    style_header_row(wg, 4, 11)

    for i in range(6):
        r = 5 + i
        ir = 5 + i
        fill = WHITE_FILL if i % 2 else STRIPE
        wg.cell(r, 1, f"=Инструменты!B{ir}")
        wg.cell(r, 2, f"=Инструменты!D{ir}")
        wg.cell(r, 3, f"=Инструменты!E{ir}")
        wg.cell(r, 4, f"=Инструменты!H{ir}")
        wg.cell(r, 5, f"=SUMIF({plan}!B:B,A{r},{plan}!F:F)")
        wg.cell(r, 6, f"=SUMIF({plan}!B:B,A{r},{plan}!J:J)")
        wg.cell(r, 7, f"=SUMIF({plan}!B:B,A{r},{plan}!L:L)")
        wg.cell(r, 8, f"=SUMIF({plan}!B:B,A{r},{plan}!M:M)")
        wg.cell(r, 9, f"=IF(H{r}=0,0,E{r}/H{r})")
        wg.cell(r, 10, f"=IF(G{r}=0,0,E{r}/G{r})")
        wg.cell(r, 11, f"=IF(G$11=0,0,G{r}/G$11)")
        for c in range(1, 12):
            apply_body(wg.cell(r, c), fill=fill, align=LEFT if c in (1, 3) else CENTER)
        wg.cell(r, 4).number_format = "0.0%"
        wg.cell(r, 5).number_format = '#,##0'
        wg.cell(r, 6).number_format = '#,##0'
        wg.cell(r, 7).number_format = '#,##0'
        wg.cell(r, 8).number_format = '#,##0'
        wg.cell(r, 9).number_format = "0.0%"
        wg.cell(r, 10).number_format = '#,##0.0'
        wg.cell(r, 11).number_format = "0.0%"

    wg.cell(11, 1, "Итого год")
    wg.cell(11, 4, "=SUM(D5:D10)")
    wg.cell(11, 5, "=SUM(E5:E10)")
    wg.cell(11, 6, "=SUM(F5:F10)")
    wg.cell(11, 7, "=SUM(G5:G10)")
    wg.cell(11, 8, "=SUM(H5:H10)")
    wg.cell(11, 9, "=IF(H11=0,0,E11/H11)")
    wg.cell(11, 10, "=IF(G11=0,0,E11/G11)")
    wg.cell(11, 11, "=SUM(K5:K10)")
    for c in range(1, 12):
        apply_body(wg.cell(11, c), fill=TOTAL_FILL)
        wg.cell(11, c).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    wg["D11"].number_format = "0.0%"
    wg["E11"].number_format = '#,##0'
    wg["F11"].number_format = '#,##0'
    wg["G11"].number_format = '#,##0'
    wg["H11"].number_format = '#,##0'
    wg["I11"].number_format = "0.0%"
    wg["J11"].number_format = '#,##0.0'
    wg["K11"].number_format = "0.0%"

    wg["A13"] = "Сверка с сеткой года"
    wg["A13"].font = FONT_H2
    wg["A14"] = "Расход (сводка)"
    wg["B14"] = "=E11"
    wg["C14"] = "Расход (сетка)"
    wg["D14"] = f"={plan}!F{total_row}"
    wg["E14"] = "Расхождение"
    wg["F14"] = "=B14-D14"
    wg["A15"] = "Выручка (сводка)"
    wg["B15"] = "=H11"
    wg["C15"] = "Выручка (сетка)"
    wg["D15"] = f"={plan}!M{total_row}"
    wg["E15"] = "Расхождение"
    wg["F15"] = "=B15-D15"
    wg["A16"] = "ДРР в коридоре 15–25%"
    wg["B16"] = '=IF(AND(I11>=0.15,I11<=0.25),"Да","Нет — проверьте бенчи")'
    for r in range(14, 17):
        for c in range(1, 7):
            apply_body(wg.cell(r, c), fill=GREEN)
    wg["B14"].number_format = '#,##0'
    wg["D14"].number_format = '#,##0'
    wg["F14"].number_format = '#,##0'
    wg["B15"].number_format = '#,##0'
    wg["D15"].number_format = '#,##0'
    wg["F15"].number_format = '#,##0'

    wg.conditional_formatting.add(
        "I5:I10",
        ColorScaleRule(
            start_type="min", start_color="27AE60",
            mid_type="percentile", mid_value=50, mid_color="F7DC6F",
            end_type="max", end_color="E74C3C",
        ),
    )

    pie_src = BarChart()
    pie_src.type = "bar"
    pie_src.grouping = "clustered"
    pie_src.title = "Покупки за год по кампаниям"
    data = Reference(wg, min_col=7, min_row=4, max_row=10)
    cats = Reference(wg, min_col=1, min_row=5, max_row=10)
    pie_src.add_data(data, titles_from_data=True)
    pie_src.set_categories(cats)
    pie_src.style = 10
    pie_src.legend = None
    pie_src.width = 18
    pie_src.height = 8
    wg.add_chart(pie_src, "A26")

    footer(
        wg, 18,
        FOOTER_YEAR_SUMMARY,
        11,
        body_rows=4,
    )
    set_widths(
        wg,
        {
            "A": 34, "B": 14, "C": 28, "D": 12, "E": 16, "F": 14,
            "G": 14, "H": 16, "I": 10, "J": 12, "K": 16,
        },
    )
    wg.auto_filter.ref = "A4:K10"

    for sheet in (ws, wi, wy, wm, wg):
        sheet.page_setup.horizontalCentered = True
        sheet.sheet_view.zoomScale = 100
        sheet.oddHeader.center.text = "&B ECCO Kids · медиаплан Директ"
        sheet.oddFooter.left.text = "&D"
        sheet.oddFooter.right.text = "стр. &P / &N"

    wb.properties.title = "ECCO Kids — медиаплан Яндекс Директ (формулы)"
    wb.properties.creator = "Polina mediaplan builder"
    wb.properties.description = (
        "Formula-driven 12-month Yandex Direct mediaplan for ECCO kids. "
        "Yellow cells are model inputs."
    )

    saved = save_workbook(wb)
    title_row, header_row, last_data_row = verify_justification(saved)
    stats = summarize_seasonality()
    print(
        "saved",
        saved,
        "total_row",
        total_row,
        "last_data",
        last_data,
        "split_title",
        title_row,
        "split_header",
        header_row,
        "split_last",
        last_data_row,
        "weight_table",
        stats["weight_table"],
        "sep_spend",
        stats["sep_spend"],
        "jul_spend",
        stats["jul_spend"],
        "year_spend",
        stats["year_spend"],
        "year_drr_pct",
        round(stats["year_drr_pct"], 1),
    )


if __name__ == "__main__":
    build()
